"""
web.py — Flask application (runs locally and on Vercel serverless).
"""

import csv
import io
import logging
import os

from flask import Flask, jsonify, render_template, request, send_file

from data_loader import load_data, reload, get_all_generic_names, get_all_molecule_names
from matcher import search, batch_search, invalidate_cache

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
logger = logging.getLogger(__name__)

# Use absolute paths so Flask finds templates/static when imported from api/
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static"),
)
app.secret_key = os.environ.get("SECRET_KEY", "pharma-search-key")
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# ── Pre-load data on startup ──────────────────────────────────────────────────
logger.info("Pre-loading datasets…")
try:
    load_data()
    logger.info("Datasets ready.")
except Exception as exc:
    logger.error("Failed to pre-load: %s", exc)


# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/search", methods=["POST"])
def search_route():
    body  = request.get_json(silent=True) or {}
    query = (body.get("query") or "").strip()
    if not query:
        return jsonify({"error": "Empty query"}), 400
    results = search(query, max_results=50)
    data    = [r.to_dict() for r in results]
    return jsonify({"query": query, "count": len(data), "results": data})


@app.route("/batch", methods=["POST"])
def batch_route():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    fname = file.filename.lower()
    try:
        if fname.endswith(".csv") or fname.endswith(".txt"):
            content = file.read().decode("utf-8", errors="replace")
            lines   = [ln.strip() for ln in content.splitlines() if ln.strip()]
            if len(lines) == 1 and "," in lines[0]:
                queries = [q.strip() for q in lines[0].split(",") if q.strip()]
            else:
                queries = lines

        elif fname.endswith(".xlsx") or fname.endswith(".xls"):
            # Use openpyxl directly — no pandas required on Vercel
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            queries = []
            for row in ws.iter_rows(values_only=True):
                val = row[0] if row else None
                if val and str(val).strip():
                    queries.append(str(val).strip())
            wb.close()

        else:
            return jsonify({"error": "Unsupported file. Use CSV, TXT, or XLSX."}), 400

    except Exception as exc:
        logger.exception("Batch parse error")
        return jsonify({"error": f"Could not parse file: {exc}"}), 400

    if not queries:
        return jsonify({"error": "No queries found in file"}), 400

    results = batch_search(queries, max_per_query=20)
    return jsonify({"count": len(results), "queries": len(queries), "results": results})


@app.route("/autocomplete")
def autocomplete():
    q = (request.args.get("q") or "").strip().lower()
    if len(q) < 2:
        return jsonify([])
    combined = sorted(set(get_all_generic_names() + get_all_molecule_names()))
    return jsonify([n for n in combined if q in n.lower()][:20])


@app.route("/stats")
def stats():
    drugs, products = load_data()
    return jsonify({"drug_entries": len(drugs), "product_entries": len(products)})


@app.route("/reload", methods=["POST"])
def reload_data():
    try:
        drugs, products = reload()
        invalidate_cache()
        return jsonify({"status": "ok", "drug_entries": len(drugs), "product_entries": len(products)})
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


# ── Local dev entry point ─────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=False, port=5000, use_reloader=False, threaded=True)
